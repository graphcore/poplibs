#!/usr/bin/env python

"""
Command line script for tabulating alexnet benchmark results.

The script must be run inside the neural nets directory. It runs the alexnet
binary multiple times with different options and writes the a summary of the
results to an Excel spreadsheet called 'alexnet.xlsx'.
"""

from __future__ import print_function
import itertools
from functools import partial
from openpyxl import Workbook
from openpyxl.styles import Font
import argparse
import re
import subprocess
import sys
from benchmark_report import create_report
import os

fields = [
    ('Number of vertices',
     'Number of vertices\s*:\s*(?P<value>[0-9.]+)'),
    ('Number of edges',
     'Number of edges\s*:\s*(?P<value>[0-9.]+)'),
    ('Vertex data',
     'Vertex data\s*:\s*(?P<value>[0-9.]+)'),
    ('Tensor data',
     'Tensor data\s*:\s*(?P<value>[0-9.]+)'),
    ('Pipelined output copies',
     'Pipelined output copies\s*:\s*(?P<value>[0-9.]+)'),
    ('In edge pointers',
     'In edge pointers\s*:\s*(?P<value>[0-9.]+)'),
    ('Message memory',
     'Message memory\s*:\s*(?P<value>[0-9.]+)'),
    ('Run instructions',
     'Run instructions\s*:\s*(?P<value>[0-9.]+)'),
    ('Num tiles exchanging',
     'Num tiles exchanging\s*:\s*(?P<value>[0-9.]+)'),
    ('Num tiles computing',
     'Num tiles computing\s*:\s*(?P<value>[0-9.]+)'),
    ('Exchange supervisor code',
     'Exchange supervisor code\s*:\s*(?P<value>[0-9.]+)'),
    ('Compute cycles',
     'Compute cycles\s*:\s*(?P<value>[0-9.]+)'),
    ('Global exchange cycles',
     'Global exchange cycles\s*:\s*(?P<value>[0-9.]+)'),
    ('Send',
     'Send\s*:\s*(?P<value>[0-9.]+)'),
    ('Receive mux',
     'Receive mux\s*:\s*(?P<value>[0-9.]+)'),
    ('Receive ptr',
     'Receive ptr\s*:\s*(?P<value>[0-9.]+)'),
    ('Nop',
     'Nop\s*:\s*(?P<value>[0-9.]+)'),
    ('Tile sync',
     'Tile sync\s+:\s+(?P<value>[0-9.]+)'),
    ('IPU sync',
     'IPU sync\s+:\s*(?P<value>[0-9.]+)'),
    ('Global sync',
     'Global sync\s*:\s*(?P<value>[0-9.]+)'),
    ('Exchange activity',
     'Exchange activity\s*:\s*(?P<value>[0-9.]+%)'),
    ('FLOPS',
     'Total number of FLOPs\s*:\s*(?P<value>[0-9.]+)'),
    ('Perfect cycles',
     'Perfect cycle time\s*:\s*(?P<value>[0-9.]+)')
]

aggregated_fields = [
    ('Parameters',
     '        Params: (?P<value>[0-9.]+)')
]

param_info = {'--ipus': {'desc': 'Num IPUs', 'default': '1'},
              '--graph-reuse': {'desc':'Reuse graphs', 'default': '1'},
              '--tiles-per-ipu': {'desc':'Tiles per IPU', 'default': '1216'},
              '--ipu-exchange-bandwidth': {'desc':'IPU exchange bandwidth',
                                            'default':'4'},
              '--bytes-per-tile': {'desc':'Bytes per tile', 'default': '262144'},
              '--data-path-width': {'desc':'Datapath width', 'default': '64'},
              '--train':{'desc':'Training', 'default': 0}}


def run(prog, params):
    cmd = ['benchmarks/{}'.format(prog)]
    cmd += ['%s=%s'%(p,v) for (p,v) in params]
    return subprocess.check_output(cmd).split('\n')


def write_headings(ws, headings, row):
    bold_font = Font(bold=True)

    for x, (heading, _) in enumerate(headings):
        column = x + 1
        ws.cell(row=row, column=column).value = heading
        ws.cell(row=row, column=column).font = bold_font

def write_data(ws, row, headings, params, data):
    for i, (heading, f) in enumerate(headings):
        column = i + 1
        f(heading, ws.cell(row=row, column=column), params, data)

def write(runs, filename):
    wb = Workbook()

    def setParamCell(heading, cell, params, data):
        cell.value = params[heading]

    def setDataCell(heading, cell, params, data, number_format=None):
        # Infomation may be missing in some cases, for example the number of
        # global exchange cycles is not reported if only 1 IPU is targeted.
        if heading in data:
            cell.value = data[heading]
        if number_format:
            cell.number_format = number_format

    def setTotalCell(firstOffset, lastOffset, heading, cell, params, data):
        first = cell.offset(column=firstOffset)
        last = cell.offset(column=lastOffset)
        cell.value = '=SUM({}:{})'.format(first.coordinate, last.coordinate)

    def completeParams(params):
        params = dict(params)
        complete_params = {}
        for (param, info) in param_info.iteritems():
            try:
                value = params[param]
            except KeyError:
                value = info['default']
            complete_params[info['desc']] =  value
        return complete_params


    headings1 = [
        ('Number of vertices', setDataCell),
        ('Number of edges', setDataCell),
        ('Vertex data', setDataCell),
        ('Tensor data', setDataCell),
        ('Pipelined output copies', setDataCell),
        ('In edge pointers', setDataCell),
        ('Message memory', setDataCell),
        ('Run instructions', setDataCell),
        ('Exchange supervisor code', setDataCell),
        ('Total memory usage', partial(setTotalCell, -7, -1)),
    ]

    headings2 = [
        ('Compute cycles', setDataCell),
        ('Global exchange cycles', setDataCell),
        ('Send', setDataCell),
        ('Receive mux', setDataCell),
        ('Receive ptr', setDataCell),
        ('Nop', setDataCell),
        ('Tile sync', setDataCell),
        ('IPU sync', setDataCell),
        ('Global sync', setDataCell),
        ('Total cycle count', partial(setTotalCell, -9, -1)),
        ('Exchange activity', partial(setDataCell, number_format='0.0%')),
        ('Exchange supervisor code', setDataCell),
        ('Message memory', setDataCell),
        ('Run instructions', setDataCell),
        ('Num tiles computing', setDataCell),
        ('Num tiles exchanging', setDataCell),
    ]

    summary_headings = []
    for (param, info) in param_info.iteritems():
        summary_headings.append((info['desc'], setParamCell))
    summary_headings += headings1 + headings2

    # Create the summary sheet
    ws = wb.active
    ws.title = 'Summary'
    write_headings(ws, summary_headings, 1)
    for y, (params, logname, layer_data) in enumerate(runs):
        data = layer_data[0]
        row = y + 2
        write_data(ws, row, summary_headings, completeParams(params), data)

    layer_headings = [('Layer ID', setDataCell)] + headings2

    # Write a sheet for each run
    for y, (params, logname, layer_data) in enumerate(runs):
        cparams = completeParams(params)
        description = ','.join(["{0}={1}".format(*p) for p in cparams.iteritems()])
        ws.cell(row=1, column=1).value = description
        ws.cell(row=1, column=1).font = Font(bold=True)
        title = "Scenario %d" % y
        ws = wb.create_sheet(title = title)
        write_headings(ws, layer_headings, 2)

        for y, data in enumerate(layer_data[1:]):
            row = y + 3
            write_data(ws, row, layer_headings, cparams, data)

    # Save the file
    wb.save(filename)

def parse(lines):
    # Build up a list of data dictionaries for each layer (computeset) in the
    # program. The first element of this list will be the totals for the entire
    # program.
    layer_data = []
    data0 = dict()
    data = data0
    found_start = False
    for line in lines:
        line = line.rstrip()
        m = re.match('(?P<name>.*) \(\d+ execution.*', line)
        if m:
            layer_data.append(data)
            data = dict()
            data['Layer ID'] = m.group('name')
        for (name, expr) in fields:
            # The report starts with an summary of the overall cycle count for
            # the whole application. This is followed by per IPU cycle counts
            # (if there are multiple IPUs) and per compute set cycle counts.
            # As a result the same field may appear multiple times in the
            # output. We only look at the first occourance of the field since
            # that corresponds to the overall cycle count.
            if name in data:
                continue
            m = re.match(expr, line)
            if m:
                str = m.group('value')
                if str.endswith('%'):
                    data[name] = float(str[:-1]) / 100
                else:
                    data[name] = float(str)

        for (name, expr) in aggregated_fields:
            # Aggregated fields get summed into the first dictionary.
            m = re.match(expr, line)
            if m:
                if name not in data:
                    data[name] =  0
                data[name] += float(m.group('value'))

    layer_data.append(data)

    return layer_data


def parse_benchmark_spec(spec):
    xs = spec.split(' ')
    prog = xs[0]
    xs = xs[1:]
    param_space = []
    param_names = []
    for x in xs:
        m = re.match('(.*)=\{?([^\}]*)\}?',x)
        if not m:
            raise Exception("Cannot parse argument: %s"%x)
        argname = m.groups(0)[0]
        vals = m.groups(0)[1].split(',')
        param_space.append((argname, vals))
        param_names.append(argname)
    scenarios = []
    for param_set in \
            itertools.product(*(options for _, options in param_space)):
        params = zip(param_names, param_set)
        logname_components = [prog]
        for (p, v) in params:
            param_str = param_info[p]['desc'].replace(' ','_')
            param_str += '_' + str(v)
            logname_components.append(param_str)
        logname = '_'.join(logname_components)
        scenarios.append((logname, params))

    return prog, scenarios

def run_benchmark(run_name, spec, args, runs):
    prog, scenarios = parse_benchmark_spec(spec)
    for logname, params in scenarios:
        if run_name:
            logname += '_' + run_name
        logname += ".log"
        if args.use_logs and os.path.exists(logname):
            print("Reading " + logname)
            log = open(logname).readlines()
        else:
            print("Creating " + logname)
            log = run(prog, params)
            with open(logname, "w") as f:
                for line in log:
                    f.write(line + "\n")
        data = parse(log)
        if prog not in runs:
            runs[prog] = []
        runs[prog].append((params, logname, data))

def write_spreadsheets(runs, run_name):
    for (prog, rs) in runs.iteritems():
        sheet_name = prog
        if run_name:
            sheet_name += '_' + run_name
        sheet_name += '.xlsx'
        print("Writing %s" % sheet_name)
        write(rs, sheet_name)
    return runs

def main():
    large_tile_opts = '--tiles-per-ipu=608 --ipu-exchange-bandwidth=8 --bytes-per-tile=524288 --data-path-width=128'
    xlarge_tile_opts = '--tiles-per-ipu=304 --ipu-exchange-bandwidth=16 --bytes-per-tile=1048576 --data-path-width=256'
    benchmarks = ['alexnet --ipus={1,2}',
                  'resnet34',
                  'resnet50']

    training_benchmarks = ['alexnet --train=1',
                           'resnet34 --train=1',
                           'resnet50 --train=1']

    arch_explore_benchmarks = ['alexnet ' + large_tile_opts,
                               'alexnet ' + xlarge_tile_opts,
                               'resnet34 ' + large_tile_opts,
                               'resnet34 ' + xlarge_tile_opts,
                               'resnet50 ' + large_tile_opts,
                               'resnet50 ' + xlarge_tile_opts,
                               'alexnet --ipu-exchange-bandwidth=8',
                               'resnet34 --ipu-exchange-bandwidth=8',
                               'resnet50 --ipu-exchange-bandwidth=8',
                               'alexnet --ipu-exchange-bandwidth=16',
                               'resnet34 --ipu-exchange-bandwidth=16',
                               'resnet50 --ipu-exchange-bandwidth=16',
                               'alexnet --ipu-exchange-bandwidth=8 --tiles-per-ipu=1024',
                               'resnet34 --ipu-exchange-bandwidth=8 --tiles-per-ipu=1024',
                               'resnet50 --ipu-exchange-bandwidth=8 --tiles-per-ipu=1024',
                               'alexnet --ipu-exchange-bandwidth=16 --tiles-per-ipu=1024',
                               'resnet34 --ipu-exchange-bandwidth=16 --tiles-per-ipu=1024',
                               'resnet50 --ipu-exchange-bandwidth=16 --tiles-per-ipu=1024'
    ]

    parser = argparse.ArgumentParser(description='Run neural net benchmarks')
    parser.add_argument('--uselogs', dest='use_logs', action='store_true',
                        help='Do not run programs. Just use previous logs.')
    parser.add_argument('--test-reuse', dest='test_reuse', action='store_true',
                        help='Test graph reuse option in resnet benchmarks.')
    parser.add_argument('--arch-explore', dest='arch_explore', action='store_true',
                        help='Explore a number of architectural configurations.')
    parser.add_argument('--training', dest='training', action='store_true',
                        help='Benchmark training.')
    parser.add_argument('--report', dest='create_report', action='store_true',
                        help='Create a overall report on the benchmarks.')
    parser.add_argument('--name', dest='run_name', default="", type=str,
                        help='Name for this benchmark run'
                               + ' (will be appended to logfile names).')

    args = parser.parse_args()
    if args.test_reuse:
        benchmarks.append['resnet34 --graph-reuse=0']
        benchmarks.append['resnet50 --graph-reuse=0']

    if args.arch_explore:
        benchmarks += arch_explore_benchmarks

    if args.training:
        benchmarks += training_benchmarks

    runs = {}
    for benchmark in benchmarks:
        run_benchmark(args.run_name, benchmark, args, runs)
    write_spreadsheets(runs, args.run_name)
    if args.create_report:
        report_file = "benchmark_report"
        if args.run_name:
            report_file += "_" + args.run_name
        report_file += ".csv"
        create_report(runs, report_file, param_info, args.arch_explore,
                      args.training)
    return 0


if __name__ == "__main__":
    sys.exit(main())

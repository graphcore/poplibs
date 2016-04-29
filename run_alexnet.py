#!/usr/bin/env python

"""
Command line script for tabulating alexnet benchmark results.

The script must be run inside the neural nets directory. It runs the alexnet
binary multiple times with different options and writes the a summary of the
results to an Excel spreadsheet called 'alexnet.xlsx'.
"""

from __future__ import print_function
import itertools
from functools import partial
from openpyxl import Workbook
from openpyxl.styles import Font
import re
import subprocess
import sys

fields = [
    ('Vertex data',
     '    Vertex data: (?P<value>[0-9.]+)'),
    ('Tensor data',
     '    Tensor data: (?P<value>[0-9.]+)'),
    ('Pipelined output copies',
     '    Pipelined output copies: (?P<value>[0-9.]+)'),
    ('In edge pointers',
     '    In edge pointers: (?P<value>[0-9.]+)'),
    ('Message memory',
     '    Message memory: (?P<value>[0-9.]+)'),
    ('Run instructions',
     '    Run instructions: (?P<value>[0-9.]+)'),
    ('Exchange supervisor code',
     '    Exchange supervisor code: (?P<value>[0-9.]+)'),
    ('Compute cycles',
     '  Compute cycles: (?P<value>[0-9.]+)'),
    ('Global exchange cycles',
     '  Global exchange cycles: (?P<value>[0-9.]+)'),
    ('Send',
     '    Send: (?P<value>[0-9.]+)'),
    ('Receive mux',
     '    Receive mux: (?P<value>[0-9.]+)'),
    ('Receive ptr',
     '    Receive ptr: (?P<value>[0-9.]+)'),
    ('Nop',
     '    Nop: (?P<value>[0-9.]+)'),
    ('Tile sync',
     '    Tile sync: (?P<value>[0-9.]+)'),
    ('IPU sync',
     '    IPU sync: (?P<value>[0-9.]+)'),
    ('Global sync',
     '    Global sync: (?P<value>[0-9.]+)')
]


def run(params):
    cmd = ['bin/alexnet']
    for name, value in params.iteritems():
        if name == 'Exchange Type':
            pass
        elif name == 'Num IPUs':
            cmd.extend(['--ipus', str(value)])
        else:
            raise Exception
    return subprocess.check_output(cmd).split('\n')


def write_headings(ws, headings):
    bold_font = Font(bold=True)

    for x, (heading, _) in enumerate(headings):
        column = x + 1
        ws.cell(row=1, column=column).value = heading
        ws.cell(row=1, column=column).font = bold_font

def write_data(ws, row, headings, params, data):
    for i, (heading, f) in enumerate(headings):
        column = i + 1
        f(heading, ws.cell(row=row, column=column), params, data)

def write(runs, filename):
    wb = Workbook()

    def setParamCell(heading, cell, params, data):
        cell.value = params[heading]

    def setDataCell(heading, cell, params, data):
        # Infomation may be missing in some cases, for example the number of
        # global exchange cycles is not reported if only 1 IPU is targeted.
        if heading in data:
            cell.value = data[heading]

    def setTotalCell(firstOffset, lastOffset, heading, cell, params, data):
        first = cell.offset(column=firstOffset)
        last = cell.offset(column=lastOffset)
        cell.value = '=SUM({}:{})'.format(first.coordinate, last.coordinate)

    headings1 = [
        ('Num IPUs', setParamCell),
        ('Vertex data', setDataCell),
        ('Tensor data', setDataCell),
        ('Pipelined output copies', setDataCell),
        ('In edge pointers', setDataCell),
        ('Message memory', setDataCell),
        ('Run instructions', setDataCell),
        ('Exchange supervisor code', setDataCell),
        ('Total memory usage', partial(setTotalCell, -7, -1)),
    ]

    headings2 = [
        ('Compute cycles', setDataCell),
        ('Global exchange cycles', setDataCell),
        ('Send', setDataCell),
        ('Receive mux', setDataCell),
        ('Receive ptr', setDataCell),
        ('Nop', setDataCell),
        ('Tile sync', setDataCell),
        ('IPU sync', setDataCell),
        ('Global sync', setDataCell),
        ('Total cycle count', partial(setTotalCell, -9, -1))
    ]

    summary_headings = headings1 + headings2

    # Create the summary sheet
    ws = wb.active
    ws.title = 'Summary'
    write_headings(ws, summary_headings)
    for y, (params, layer_data) in enumerate(runs):
        data = layer_data[0]
        row = y + 2
        write_data(ws, row, summary_headings, params, data)

    layer_headings = [('Layer ID', setDataCell)] + headings2

    # Write a sheet for each run
    for y, (params, layer_data) in enumerate(runs):
        title = ','.join(["{0}={1}".format(*p) for p in params.iteritems()])
        ws = wb.create_sheet(title = title)
        write_headings(ws, layer_headings)

        for y, data in enumerate(layer_data[1:]):
            row = y + 2
            write_data(ws, row, layer_headings, params, data)


    # Save the file
    wb.save(filename)


def parse(lines):
    # Build up a list of data dictionaries for each layer (computeset) in the
    # program. The first element of this list will be the totals for the entire
    # program.
    layer_data = []
    data = dict()
    found_start = False
    for line in lines:
        line = line.rstrip()
        m = re.match('(?P<name>.*) \(\d+ execution.*', line)
        if m:
            layer_data.append(data)
            data = dict()
            data['Layer ID'] = m.group('name')
        for (name, expr) in fields:
            # The report starts with an summary of the overall cycle count for
            # the whole application. This is followed by per IPU cycle counts
            # (if there are multiple IPUs) and per compute set cycle counts.
            # As a result the same field may appear multiple times in the
            # output. We only look at the first occourance of the field since
            # that corresponds to the overall cycle count.
            if name in data:
                continue
            m = re.match(expr, line)
            if m:
                data[name] = float(m.group('value'))

    layer_data.append(data)

    return layer_data


def main():
    param_space = [
        ('Num IPUs', [1, 2])
    ]
    param_names = [name for name, _ in param_space]
    runs = []
    for param_set in \
            itertools.product(*(options for _, options in param_space)):
        params = dict(zip(param_names, param_set))
        log = run(params)
        data = parse(log)
        runs.append((params, data))

    write(runs, 'alexnet.xlsx')
    return 0


if __name__ == "__main__":
    sys.exit(main())
